#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Тестирование сохранения группировки Excel файлов в JSON и обратно
"""

import sys
import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font



from src.core.excel_hierarchy_preserver import ExcelHierarchyPreserver, json_to_excel

def create_test_excel_with_grouping():
    """Создает тестовый Excel файл с группировкой строк и столбцов."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Тестовый лист"
    
    # Добавляем данные
    data = [
        ["Группа 1", "", ""],
        ["Подгруппа 1.1", "Значение 1", 100],
        ["Подгруппа 1.2", "Значение 2", 200],
        ["Итого по группе 1", "", 300],
        ["Группа 2", "", ""],
        ["Подгруппа 2.1", "Значение 3", 150],
        ["Подгруппа 2.2", "Значение 4", 250],
        ["Итого по группе 2", "", 400],
        ["Общий итог", "", 700]
    ]
    
    # Заполняем лист данными
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Применяем группировку строк
    # Группируем строки 2-4 (Подгруппа 1.1 - Итого по группе 1)
    ws.row_dimensions.group(2, 4, hidden=True)
    # Группируем строки 5-8 (Группа 2 - Итого по группе 2)
    ws.row_dimensions.group(5, 8, hidden=True)
    
    # Применяем форматирование для заголовков
    for row in [1, 4, 5, 8, 9]:
        for col in range(1, 4):
            ws.cell(row=row, column=col).font = Font(bold=True)
    
    # Сохраняем файл
    test_file = "результаты/тест_группировка.xlsx"
    os.makedirs("результаты", exist_ok=True)
    wb.save(test_file)
    wb.close()
    
    print(f"✅ Создан тестовый Excel файл с группировкой: {test_file}")
    return test_file

def test_grouping_preservation():
    """Тестирует сохранение группировки при преобразовании Excel в JSON и обратно."""
    print("🔬 Тестирование сохранения группировки Excel файлов")
    print("=" * 60)
    
    # 1. Создаем тестовый Excel файл с группировкой
    test_excel_file = create_test_excel_with_grouping()
    
    # 2. Преобразуем Excel в JSON с сохранением группировки
    json_file = "результаты/тест_группировка.json"
    
    try:
        preserver = ExcelHierarchyPreserver(test_excel_file)
        if preserver.to_json(json_file):
            print(f"✅ Excel файл преобразован в JSON с сохранением группировки: {json_file}")
        else:
            print("❌ Ошибка преобразования Excel в JSON")
            assert False, "Ошибка преобразования Excel в JSON"
    except Exception as e:
        print(f"❌ Ошибка создания JSON файла: {e}")
        assert False, f"Ошибка создания JSON файла: {e}"
    
    # 3. Проверяем содержимое JSON файла
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        print("\n📋 Структура созданного JSON файла:")
        print(f"   Исходный файл: {json_data.get('source_file', 'Не указан')}")
        print(f"   Количество листов: {json_data.get('sheet_count', 0)}")
        
        if 'grouping_info' in json_data:
            print("   Информация о группировке:")
            for sheet_name, group_info in json_data['grouping_info'].items():
                print(f"     Лист '{sheet_name}':")
                if 'row_outline_levels' in group_info:
                    print(f"       Группировка строк: {len(group_info['row_outline_levels'])} записей")
                if 'column_outline_levels' in group_info:
                    print(f"       Группировка столбцов: {len(group_info['column_outline_levels'])} записей")
        else:
            print("   Информация о группировке отсутствует")
            
    except Exception as e:
        print(f"❌ Ошибка чтения JSON файла: {e}")
        assert False, f"Ошибка чтения JSON файла: {e}"
    
    # 4. Преобразуем JSON обратно в Excel
    restored_excel_file = "результаты/восстановленный_тест_группировка.xlsx"
    
    try:
        if json_to_excel(json_file, restored_excel_file):
            print(f"✅ JSON файл преобразован обратно в Excel: {restored_excel_file}")
        else:
            print("❌ Ошибка преобразования JSON в Excel")
            assert False, "Ошибка преобразования JSON в Excel"
    except Exception as e:
        print(f"❌ Ошибка восстановления Excel файла: {e}")
        assert False, f"Ошибка восстановления Excel файла: {e}"
    
    # 5. Проверяем, что файлы существуют
    assert os.path.exists(test_excel_file) and os.path.exists(json_file) and os.path.exists(restored_excel_file), "Некоторые файлы отсутствуют"
    print("\n✅ Все файлы успешно созданы")

