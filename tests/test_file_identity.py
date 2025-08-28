#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Тестирование идентичности исходного и восстановленного Excel файлов
"""

import sys
import os
import json
import pandas as pd
from openpyxl import load_workbook

# Добавляем путь к корню проекта
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from src.core.excel_hierarchy_preserver import ExcelHierarchyPreserver, json_to_excel

def compare_excel_files_detailed(file1, file2):
    """Сравнивает содержимое двух Excel файлов детально."""
    try:
        # Сравниваем данные с помощью pandas
        xl1 = pd.ExcelFile(file1)
        xl2 = pd.ExcelFile(file2)
        
        # Сравниваем имена листов
        if xl1.sheet_names != xl2.sheet_names:
            print(f"❌ Имена листов различаются: {xl1.sheet_names} vs {xl2.sheet_names}")
            return False
        
        # Сравниваем содержимое каждого листа
        for sheet_name in xl1.sheet_names:
            df1 = xl1.parse(sheet_name, header=None)
            df2 = xl2.parse(sheet_name, header=None)
            
            # Сравниваем размеры
            if df1.shape != df2.shape:
                print(f"❌ Размеры листа '{sheet_name}' различаются: {df1.shape} vs {df2.shape}")
                return False
            
            # Сравниваем данные
            if not df1.equals(df2):
                print(f"❌ Данные листа '{sheet_name}' различаются")
                # Покажем различия
                try:
                    diff = df1.compare(df2)
                    if not diff.empty:
                        print(f"   Пример различий:\n{diff.head()}")
                except Exception:
                    # Если compare не работает, сравниваем поэлементно
                    for i in range(min(len(df1), len(df2))):
                        for j in range(min(len(df1.columns), len(df2.columns))):
                            if df1.iat[i, j] != df2.iat[i, j]:
                                print(f"   Различие в ячейке [{i}, {j}]: {df1.iat[i, j]} vs {df2.iat[i, j]}")
                                return False
                return False
        
        # Сравниваем группировку и другие свойства с помощью openpyxl
        wb1 = load_workbook(file1)
        wb2 = load_workbook(file2)
        
        for sheet_name in wb1.sheetnames:
            ws1 = wb1[sheet_name]
            ws2 = wb2[sheet_name]
            
            # Сравниваем уровни группировки строк
            row_groups1 = {row_num: (row_dim.outline_level, getattr(row_dim, 'hidden', False))
                          for row_num, row_dim in ws1.row_dimensions.items() 
                          if row_dim.outline_level > 0}
            row_groups2 = {row_num: (row_dim.outline_level, getattr(row_dim, 'hidden', False))
                          for row_num, row_dim in ws2.row_dimensions.items() 
                          if row_dim.outline_level > 0}
            
            if row_groups1 != row_groups2:
                print(f"❌ Группировка строк листа '{sheet_name}' различается")
                print(f"   Оригинал: {row_groups1}")
                print(f"   Восстановленный: {row_groups2}")
                wb1.close()
                wb2.close()
                return False
            
            # Сравниваем уровни группировки столбцов
            col_groups1 = {col_letter: (col_dim.outline_level, getattr(col_dim, 'hidden', False))
                          for col_letter, col_dim in ws1.column_dimensions.items() 
                          if col_dim.outline_level > 0}
            col_groups2 = {col_letter: (col_dim.outline_level, getattr(col_dim, 'hidden', False))
                          for col_letter, col_dim in ws2.column_dimensions.items() 
                          if col_dim.outline_level > 0}
            
            if col_groups1 != col_groups2:
                print(f"❌ Группировка столбцов листа '{sheet_name}' различается")
                print(f"   Оригинал: {col_groups1}")
                print(f"   Восстановленный: {col_groups2}")
                wb1.close()
                wb2.close()
                return False
            
            # Сравниваем объединенные ячейки
            merged1 = set(str(rng) for rng in ws1.merged_cells.ranges)
            merged2 = set(str(rng) for rng in ws2.merged_cells.ranges)
            
            if merged1 != merged2:
                print(f"❌ Объединенные ячейки листа '{sheet_name}' различаются")
                print(f"   Оригинал: {merged1}")
                print(f"   Восстановленный: {merged2}")
                wb1.close()
                wb2.close()
                return False
        
        wb1.close()
        wb2.close()
        
        return True
    except Exception as e:
        print(f"❌ Ошибка сравнения файлов: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_file_identity():
    """Тестирование идентичности исходного и восстановленного Excel файлов."""
    print("🔍 Тестирование идентичности исходного и восстановленного Excel файлов")
    print("=" * 70)
    
    # Путь к тестовому файлу
    test_excel_file = "результаты/b3ded820-4385-4a42-abc7-75dc0756d335_6ba7ac50-fb1a-4bea-9489-265bdfdcb8d1_вся номенклатура за период с 15.07.25 по 21.07.25.xlsx"
    
    # Проверяем существование тестового файла
    if not os.path.exists(test_excel_file):
        print(f"❌ Тестовый файл не найден: {test_excel_file}")
        return False
    
    print(f"📂 Работаем с файлом: {test_excel_file}")
    
    # Преобразуем Excel в JSON
    json_file = "результаты/test_identity.json"
    print("\n🔄 Преобразование Excel → JSON...")
    
    try:
        preserver = ExcelHierarchyPreserver(test_excel_file)
        if preserver.to_json(json_file):
            print(f"✅ Успешно! JSON файл сохранен: {json_file}")
        else:
            print("❌ Ошибка преобразования Excel → JSON")
            return False
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        return False
    
    # Преобразуем JSON обратно в Excel
    restored_excel_file = "результаты/test_identity_restored.xlsx"
    print("\n🔄 Преобразование JSON → Excel...")
    
    try:
        if json_to_excel(json_file, restored_excel_file):
            print(f"✅ Успешно! Восстановленный Excel файл: {restored_excel_file}")
        else:
            print("❌ Ошибка восстановления")
            return False
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        return False
    
    # Сравниваем содержимое файлов
    print("\n🔍 Сравнение содержимого файлов...")
    if compare_excel_files_detailed(test_excel_file, restored_excel_file):
        print("\n🎉 Содержимое файлов идентично!")
        return True
    else:
        print("\n❌ Содержимое файлов различается.")
        return False

def main():
    """Основная функция тестирования."""
    try:
        success = test_file_identity()
        if success:
            print("\n✅ Тест идентичности файлов пройден успешно!")
            print("   Исходный и восстановленный Excel файлы идентичны по содержимому и структуре.")
        else:
            print("\n❌ Тест идентичности файлов провален!")
            print("   Исходный и восстановленный Excel файлы различаются.")
    except Exception as e:
        print(f"\n💥 Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    main()