#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Финальный тест: полный цикл сохранения и восстановления Excel файлов с группировкой
"""

import sys
import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font



from src.core.excel_hierarchy_preserver import ExcelHierarchyPreserver, json_to_excel

def create_complex_excel_with_grouping():
    """Создает сложный Excel файл с многоуровневой группировкой."""
    wb = Workbook()
    
    # Создаем первый лист
    ws1 = wb.active
    ws1.title = "Продажи"
    
    # Добавляем данные с группировкой
    sales_data = [
        ["Год 2025", "", "", ""],
        ["Квартал 1", "", "", ""],
        ["Январь", "Продукт A", 100, 15000],
        ["Январь", "Продукт B", 200, 25000],
        ["Итого за Январь", "", 300, 40000],
        ["Февраль", "Продукт A", 120, 18000],
        ["Февраль", "Продукт B", 180, 22000],
        ["Итого за Февраль", "", 300, 40000],
        ["Март", "Продукт A", 110, 16500],
        ["Март", "Продукт B", 190, 23500],
        ["Итого за Март", "", 300, 40000],
        ["Итого за Квартал 1", "", 900, 120000],
        ["Квартал 2", "", "", ""],
        ["Апрель", "Продукт A", 130, 19500],
        ["Апрель", "Продукт B", 170, 21000],
        ["Итого за Апрель", "", 300, 40500],
        ["Май", "Продукт A", 140, 21000],
        ["Май", "Продукт B", 160, 20000],
        ["Итого за Май", "", 300, 41000],
        ["Июнь", "Продукт A", 150, 22500],
        ["Июнь", "Продукт B", 150, 19500],
        ["Итого за Июнь", "", 300, 42000],
        ["Итого за Квартал 2", "", 900, 123500],
        ["Итого за Год 2025", "", 1800, 243500]
    ]
    
    # Заполняем первый лист
    for row_idx, row_data in enumerate(sales_data, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Применяем группировку строк на первом листе
    # Группируем строки по кварталам
    ws1.row_dimensions.group(2, 12, hidden=True)   # Квартал 1
    ws1.row_dimensions.group(13, 23, hidden=True)  # Квартал 2
    
    # Группируем строки внутри кварталов по месяцам
    ws1.row_dimensions.group(3, 5, hidden=True)    # Январь
    ws1.row_dimensions.group(6, 8, hidden=True)    # Февраль
    ws1.row_dimensions.group(9, 11, hidden=True)   # Март
    ws1.row_dimensions.group(14, 16, hidden=True)  # Апрель
    ws1.row_dimensions.group(17, 19, hidden=True)  # Май
    ws1.row_dimensions.group(20, 22, hidden=True)  # Июнь
    
    # Создаем второй лист
    ws2 = wb.create_sheet("Закупки")
    
    # Добавляем данные со своей структурой
    purchase_data = [
        ["Поставщик X", "", ""],
        ["Продукт A", 500, 75000],
        ["Продукт B", 300, 37500],
        ["Итого от Поставщика X", 800, 112500],
        ["Поставщик Y", "", ""],
        ["Продукт A", 400, 60000],
        ["Продукт B", 400, 50000],
        ["Итого от Поставщика Y", 800, 110000],
        ["Общий итог", 1600, 222500]
    ]
    
    # Заполняем второй лист
    for row_idx, row_data in enumerate(purchase_data, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Применяем группировку строк на втором листе
    ws2.row_dimensions.group(1, 4, hidden=True)  # Поставщик X
    ws2.row_dimensions.group(5, 8, hidden=True)  # Поставщик Y
    
    # Применяем форматирование
    for ws in [ws1, ws2]:
        # Жирный шрифт для заголовков
        for row in [1, 2, 13, 24, 1, 5, 9]:
            for col in range(1, 5):
                if ws.cell(row=row, column=col).value:
                    ws.cell(row=row, column=col).font = Font(bold=True)
    
    # Сохраняем файл
    test_file = "результаты/сложный_тест_группировка.xlsx"
    os.makedirs("результаты", exist_ok=True)
    wb.save(test_file)
    wb.close()
    
    print(f"✅ Создан сложный Excel файл с группировкой: {test_file}")
    return test_file

def test_full_cycle():
    """Полный цикл тестирования: Excel → JSON → Excel с сохранением группировки."""
    print("🔬 Полный цикл тестирования сохранения группировки Excel файлов")
    print("=" * 70)
    
    # 1. Создаем сложный Excel файл с группировкой
    original_excel = create_complex_excel_with_grouping()
    
    # 2. Преобразуем Excel в JSON с сохранением группировки
    json_file = "результаты/сложный_тест_группировка.json"
    
    try:
        preserver = ExcelHierarchyPreserver(original_excel)
        if preserver.to_json(json_file):
            print(f"✅ Excel файл преобразован в JSON с сохранением группировки: {json_file}")
        else:
            assert False, "Ошибка преобразования Excel в JSON"
    except Exception as e:
        assert False, f"Ошибка создания JSON файла: {e}"
    
    # 3. Проверяем содержимое JSON файла
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        print("\n📋 Структура созданного JSON файла:")
        print(f"   Исходный файл: {json_data.get('source_file', 'Не указан')}")
        print(f"   Количество листов: {json_data.get('sheet_count', 0)}")
        
        if 'grouping_info' in json_data:
            print("   Информация о группировке:")
            for sheet_name, group_info in json_data['grouping_info'].items():
                print(f"     Лист '{sheet_name}':")
                if 'row_outline_levels' in group_info:
                    row_groups = len([k for k, v in group_info['row_outline_levels'].items() if v != 0])
                    print(f"       Группировка строк: {row_groups} групп")
                if 'column_outline_levels' in group_info:
                    col_groups = len([k for k, v in group_info['column_outline_levels'].items() if v != 0])
                    print(f"       Группировка столбцов: {col_groups} групп")
        else:
            print("   Информация о группировке отсутствует")
            
    except Exception as e:
        assert False, f"Ошибка чтения JSON файла: {e}"
    
    # 4. Преобразуем JSON обратно в Excel
    restored_excel = "результаты/восстановленный_сложный_тест_группировка.xlsx"
    
    try:
        if json_to_excel(json_file, restored_excel):
            print(f"✅ JSON файл преобразован обратно в Excel: {restored_excel}")
        else:
            assert False, "Ошибка преобразования JSON в Excel"
    except Exception as e:
        assert False, f"Ошибка восстановления Excel файла: {e}"
    
    # 5. Проверяем, что файлы существуют
    assert os.path.exists(original_excel) and os.path.exists(json_file) and os.path.exists(restored_excel), "Некоторые файлы отсутствуют"
    print("\n✅ Все файлы успешно созданы")
    
    # 6. Сравниваем структуру оригинального и восстановленного файлов
    try:
        # Открываем оба файла для сравнения
        from openpyxl import load_workbook
        
        # Оригинальный файл
        wb_orig = load_workbook(original_excel)
        ws_orig_sales = wb_orig["Продажи"]
        ws_orig_purchase = wb_orig["Закупки"]
        
        # Восстановленный файл
        wb_rest = load_workbook(restored_excel)
        ws_rest_sales = wb_rest["Продажи"]
        ws_rest_purchase = wb_rest["Закупки"]
        
        # Проверяем группировку на листе Продажи
        orig_sales_groups = {row_num: row_dim.outline_level 
                            for row_num, row_dim in ws_orig_sales.row_dimensions.items() 
                            if row_dim.outline_level > 0}
        rest_sales_groups = {row_num: row_dim.outline_level 
                            for row_num, row_dim in ws_rest_sales.row_dimensions.items() 
                            if row_dim.outline_level > 0}
        
        # Проверяем группировку на листе Закупки
        orig_purchase_groups = {row_num: row_dim.outline_level 
                              for row_num, row_dim in ws_orig_purchase.row_dimensions.items() 
                              if row_dim.outline_level > 0}
        rest_purchase_groups = {row_num: row_dim.outline_level 
                              for row_num, row_dim in ws_rest_purchase.row_dimensions.items() 
                              if row_dim.outline_level > 0}
        
        wb_orig.close()
        wb_rest.close()
        
        # Сравниваем группировки
        assert orig_sales_groups == rest_sales_groups and orig_purchase_groups == rest_purchase_groups, "Группировка строк НЕ была полностью сохранена и восстановлена"
        print("✅ Группировка строк полностью сохранена и восстановлена")

    except Exception as e:
        assert False, f"⚠️  Не удалось сравнить структуру файлов: {e}"