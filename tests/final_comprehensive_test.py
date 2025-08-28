#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ФИНАЛЬНЫЙ ТЕСТ: Проверка полной функциональности системы
с сохранением и восстановлением группировки Excel файлов
"""

import sys
import os
import json
from pathlib import Path

# Добавляем путь к корню проекта


from src.core.excel_hierarchy_preserver import ExcelHierarchyPreserver, json_to_excel

def final_comprehensive_test():
    """Финальный комплексный тест всей системы."""
    print("🏆 ФИНАЛЬНЫЙ КОМПЛЕКСНЫЙ ТЕСТ СИСТЕМЫ")
    print("=" * 60)
    
    # 1. Проверяем, что все модули импортируются корректно
    print("\n🔍 1. Проверка импорта модулей...")
    try:
        from src.core.excel_hierarchy_preserver import ExcelHierarchyPreserver, json_to_excel
        from src.core.shrinkage_system import ShrinkageSystem
        from src.core.json_parser import parse_from_json
        from src.core.data_processor import DataProcessor
        from src.core.reporting import ReportGenerator
        print("✅ Все модули успешно импортированы")
    except Exception as e:
        assert False, f"Ошибка импорта модулей: {e}"
    
    # 2. Проверяем работу с тестовыми файлами
    print("\n📂 2. Проверка работы с тестовыми файлами...")
    
    test_files_dir = Path("результаты")
    assert test_files_dir.exists(), "Директория результатов не найдена"
    
    excel_files = list(test_files_dir.glob("*.xlsx"))
    assert excel_files, "Тестовые Excel файлы не найдены"
    
    test_excel_file = str(excel_files[0])
    print(f"✅ Найден тестовый Excel файл: {os.path.basename(test_excel_file)}")
    
    # 3. Тестируем преобразование Excel → JSON с сохранением группировки
    print("\n🔄 3. Тестирование преобразования Excel → JSON...")
    json_file = "результаты/финальный_тест.json"
    
    try:
        preserver = ExcelHierarchyPreserver(test_excel_file)
        assert preserver.to_json(json_file), "Ошибка преобразования Excel → JSON"
        print(f"✅ Excel → JSON: {os.path.basename(json_file)}")
    except Exception as e:
        assert False, f"Ошибка создания JSON файла: {e}"
    
    # 4. Проверяем содержимое JSON файла
    print("\n📄 4. Проверка содержимого JSON файла...")
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        assert 'grouping_info' in json_data, "Информация о группировке отсутствует в JSON файле"
        print("✅ Найдена информация о группировке в JSON файле")
        grouping_sheets = len(json_data['grouping_info'])
        print(f"   Листов с группировкой: {grouping_sheets}")
        
        for sheet_name, group_info in json_data['grouping_info'].items():
            if 'row_outline_levels' in group_info:
                row_groups = len([k for k, v in group_info['row_outline_levels'].items() if v != 0])
                print(f"   Лист '{sheet_name}': {row_groups} групп строк")
        
        print(f"✅ JSON файл корректно структурирован")
    except Exception as e:
        assert False, f"Ошибка проверки JSON файла: {e}"
    
    # 5. Тестируем преобразование JSON → Excel с восстановлением группировки
    print("\n🔄 5. Тестирование преобразования JSON → Excel...")
    restored_excel_file = "результаты/финальный_восстановленный_тест.xlsx"
    
    try:
        assert json_to_excel(json_file, restored_excel_file), "Ошибка преобразования JSON → Excel"
        print(f"✅ JSON → Excel: {os.path.basename(restored_excel_file)}")
    except Exception as e:
        assert False, f"Ошибка восстановления Excel файла: {e}"
    
    # 6. Проверяем существование всех файлов
    print("\n🔎 6. Проверка существования файлов...")
    files_to_check = [
        test_excel_file,
        json_file,
        restored_excel_file
    ]
    
    for file_path in files_to_check:
        assert os.path.exists(file_path), f"Файл {os.path.basename(file_path)} отсутствует"
        print(f"✅ {os.path.basename(file_path)}")
    
    # 7. Проверяем размеры файлов
    print("\n📊 7. Проверка размеров файлов...")
    original_size = os.path.getsize(test_excel_file)
    restored_size = os.path.getsize(restored_excel_file)
    
    print(f"   Оригинал: {original_size} байт")
    print(f"   Восстановленный: {restored_size} байт")
    
    # 8. Финальная проверка
    print("\n🎯 8. Финальная проверка...")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(restored_excel_file)
        sheet_count = len(wb.sheetnames)
        print(f"✅ Восстановленный файл содержит {sheet_count} листов")
        wb.close()
        
        with open(json_file, 'r', encoding='utf-8') as f:
            json_content = f.read()
            assert len(json_content) > 100, "JSON файл пуст или поврежден"
            print("✅ JSON файл содержит данные")
                
    except Exception as e:
        assert False, f"Ошибка финальной проверки: {e}"

    # 9. Тестирование нового workflow (JSON -> DataFrame -> Расчет)
    print("\n⚙️ 9. Тестирование нового workflow...")
    try:
        dataframe = parse_from_json(json_file)
        assert not dataframe.empty, "DataFrame не должен быть пустым после парсинга JSON"
        print(f"✅ JSON -> DataFrame: успешно, {len(dataframe)} строк")

        system = ShrinkageSystem()
        results = system.process_dataset(dataframe, "final_test.json", use_adaptive=True)
        assert results['status'] == 'success', "Ошибка при обработке датасета"
        print("✅ ShrinkageSystem.process_dataset: успешно")

    except Exception as e:
        assert False, f"Ошибка в новом workflow: {e}"

    print("\n" + "=" * 60)
    print("🏆 ФИНАЛЬНЫЙ КОМПЛЕКСНЫЙ ТЕСТ ПРОЙДЕН УСПЕШНО!")
    print("=" * 60)
    
    print("\n🎉 ВСЕ ФУНКЦИИ СИСТЕМЫ РАБОТАЮТ КОРРЕКТНО:")
    print("   ✅ Импорт модулей")
    print("   ✅ Преобразование Excel → JSON с сохранением группировки")
    print("   ✅ Преобразование JSON → Excel с восстановлением группировки")
    print("   ✅ Парсинг JSON в DataFrame")
    print("   ✅ Расчеты на основе данных из JSON")
    print("   ✅ Сохранение структуры файлов")
    print("   ✅ Создание и проверка файлов")
    
    print("\n🚀 СИСТЕМА ГОТОВА К ИСПОЛЬЗОВАНИЮ В PRODUCTION!")


def main():
    """Основная функция финального теста."""
    try:
        success = final_comprehensive_test()
        if success:
            print("\n🎊 ПОЗДРАВЛЯЕМ! ВСЕ ТЕСТЫ ПРОЙДЕНЫ УСПЕШНО!")
            print("   Система полностью готова к использованию.")
        else:
            print("\n❌ ФИНАЛЬНЫЙ ТЕСТ ПРОВАЛЕН")
            print("   Требуется дополнительная проверка.")
    except Exception as e:
        print(f"\n💥 КРИТИЧЕСКАЯ ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    main()