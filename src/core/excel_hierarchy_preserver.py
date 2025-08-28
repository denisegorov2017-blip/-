"""
Модуль для сохранения иерархии Excel файлов и преобразования в различные форматы
Сохраняет данные, структуру листов и информацию о группировке строк/столбцов
"""

import pandas as pd
import json
import os
from typing import Dict, Any, List, Union
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class ExcelHierarchyPreserver:
    """Класс для сохранения иерархии Excel файлов и преобразования в различные форматы."""
    
    def __init__(self, excel_file_path: str):
        """
        Инициализация класса.
        
        Args:
            excel_file_path: Путь к Excel файлу
        """
        self.excel_file_path = excel_file_path
        self.data_structure = None
        self.grouping_info = None
        self.formatting_info = None  # Новая переменная для хранения информации о форматировании
        self._load_excel_structure()
    
    def _load_excel_structure(self) -> None:
        """Загружает структуру Excel файла с сохранением иерархии и группировки."""
        file_extension = Path(self.excel_file_path).suffix.lower()
        
        try:
            if file_extension in ['.xlsx', '.xlsm']:
                # Используем openpyxl для получения информации о группировке
                wb = load_workbook(self.excel_file_path, data_only=True)
                excel_file = pd.ExcelFile(self.excel_file_path, engine='openpyxl')
            elif file_extension == '.xls':
                excel_file = pd.ExcelFile(self.excel_file_path, engine='xlrd')
                wb = None  # Для .xls нет поддержки группировки в текущей реализации
            else:
                raise ValueError(f"Неподдерживаемый формат файла: {file_extension}")
            
            # Сохраняем имена листов
            self.sheet_names = excel_file.sheet_names
            
            # Загружаем данные для каждого листа
            self.data_structure = {}
            self.grouping_info = {}
            self.formatting_info = {}  # Инициализируем новую переменную
            self.sheet_dimensions = {}  # Новая переменная для хранения размеров листов
            
            for sheet_name in self.sheet_names:
                # Загружаем лист как DataFrame
                df = excel_file.parse(sheet_name=sheet_name, header=None)
                # Сохраняем как матрицу значений
                self.data_structure[sheet_name] = df.values.tolist()
                # Сохраняем размеры листа
                self.sheet_dimensions[sheet_name] = (df.shape[0], df.shape[1])
                
                # Получаем информацию о группировке (только для .xlsx/.xlsm)
                if wb and sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    self.grouping_info[sheet_name] = self._extract_grouping_info(ws)
                    # Получаем информацию о форматировании
                    self.formatting_info[sheet_name] = self._extract_formatting_info(ws)
            
            if wb:
                wb.close()
            excel_file.close()
            
        except Exception as e:
            raise ValueError(f"Ошибка загрузки Excel файла: {str(e)}")
    
    def _extract_grouping_info(self, worksheet) -> Dict[str, Any]:
        """
        Извлекает информацию о группировке строк и столбцов из листа Excel.
        
        Args:
            worksheet: Объект листа openpyxl
            
        Returns:
            Словарь с информацией о группировке
        """
        grouping_info = {
            'row_groups': [],
            'column_groups': [],
            'row_outline_levels': {},
            'column_outline_levels': {}
        }
        
        # Получаем информацию о группировке строк
        for row_num, row_dim in worksheet.row_dimensions.items():
            if hasattr(row_dim, 'outline_level') and row_dim.outline_level > 0:
                grouping_info['row_outline_levels'][str(row_num)] = row_dim.outline_level
                # Проверяем, свернута ли строка
                if hasattr(row_dim, 'hidden') and row_dim.hidden:
                    grouping_info['row_outline_levels'][str(row_num)] = -grouping_info['row_outline_levels'][str(row_num)]
        
        # Получаем информацию о группировке столбцов
        for col_letter, col_dim in worksheet.column_dimensions.items():
            if hasattr(col_dim, 'outline_level') and col_dim.outline_level > 0:
                grouping_info['column_outline_levels'][col_letter] = col_dim.outline_level
                # Проверяем, свернут ли столбец
                if hasattr(col_dim, 'hidden') and col_dim.hidden:
                    grouping_info['column_outline_levels'][col_letter] = -grouping_info['column_outline_levels'][col_letter]
        
        return grouping_info
    
    def _extract_formatting_info(self, worksheet) -> Dict[str, Any]:
        """
        Извлекает информацию о форматировании из листа Excel.
        
        Args:
            worksheet: Объект листа openpyxl
            
        Returns:
            Словарь с информацией о форматировании
        """
        formatting_info = {
            'cell_formats': {},
            'merged_cells': [],
            'sheet_properties': {}
        }
        
        # Получаем информацию о форматировании ячеек
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.font or cell.fill or cell.border or cell.alignment:
                    cell_info = {}
                    if cell.font:
                        cell_info['font'] = {
                            'name': cell.font.name,
                            'size': cell.font.size,
                            'bold': cell.font.bold,
                            'italic': cell.font.italic,
                            'color': cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
                        }
                    if cell.fill:
                        cell_info['fill'] = {
                            'patternType': cell.fill.patternType,
                            'fgColor': cell.fill.fgColor.rgb if cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb') else None
                        }
                    if cell.border:
                        cell_info['border'] = {
                            'left': cell.border.left.style if cell.border.left else None,
                            'right': cell.border.right.style if cell.border.right else None,
                            'top': cell.border.top.style if cell.border.top else None,
                            'bottom': cell.border.bottom.style if cell.border.bottom else None
                        }
                    if cell.alignment:
                        cell_info['alignment'] = {
                            'horizontal': cell.alignment.horizontal,
                            'vertical': cell.alignment.vertical
                        }
                    
                    if cell_info:
                        formatting_info['cell_formats'][f"{cell.row}_{cell.column_letter}"] = cell_info
        
        # Получаем информацию о объединенных ячейках
        for merged_range in worksheet.merged_cells.ranges:
            formatting_info['merged_cells'].append(str(merged_range))
        
        # Получаем свойства листа
        formatting_info['sheet_properties'] = {
            'sheet_state': worksheet.sheet_state,
            'page_setup': {
                'orientation': worksheet.page_setup.orientation if worksheet.page_setup else None,
                'paperSize': worksheet.page_setup.paperSize if worksheet.page_setup else None
            }
        }
        
        return formatting_info
    
    def to_json(self, output_path: str) -> bool:
        """
        Преобразует Excel файл в JSON с сохранением иерархии и группировки.
        
        Args:
            output_path: Путь для сохранения JSON файла
            
        Returns:
            True если успешно
        """
        try:
            # Создаем директорию если не существует
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Подготавливаем данные для сохранения
            json_data = {
                "source_file": self.excel_file_path,
                "sheet_count": len(self.sheet_names),
                "sheets": {},
                "grouping_info": self.grouping_info if self.grouping_info else {},
                "formatting_info": self.formatting_info if self.formatting_info else {},  # Добавляем информацию о форматировании
                "sheet_dimensions": self.sheet_dimensions if self.sheet_dimensions else {}  # Добавляем размеры листов
            }
            
            # Добавляем данные каждого листа
            for sheet_name, sheet_data in self.data_structure.items():
                json_data["sheets"][sheet_name] = sheet_data
            
            # Сохраняем в JSON файл
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, indent=2, default=str)
            
            return True
        except Exception as e:
            print(f"Ошибка сохранения в JSON: {e}")
            return False
    
    def get_structure_info(self) -> Dict[str, Any]:
        """
        Возвращает информацию о структуре Excel файла.
        
        Returns:
            Словарь с информацией о структуре
        """
        if not self.data_structure:
            return {"error": "Структура не загружена"}
        
        info = {
            "source_file": self.excel_file_path,
            "sheet_count": len(self.sheet_names),
            "sheets": {}
        }
        
        for sheet_name, sheet_data in self.data_structure.items():
            info["sheets"][sheet_name] = {
                "rows": len(sheet_data),
                "columns": len(sheet_data[0]) if sheet_data else 0
            }
        
        return info


def json_to_excel(json_path: str, output_excel_path: str) -> bool:
    """
    Преобразует JSON файл обратно в Excel с сохранением структуры и группировки.
    
    Args:
        json_path: Путь к JSON файлу
        output_excel_path: Путь для сохранения Excel файла
        
    Returns:
        True если успешно
    """
    try:
        # Читаем JSON файл
        with open(json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        # Создаем директорию если не существует
        os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
        
        # Проверяем наличие данных листов
        if "sheets" not in json_data:
            raise ValueError("JSON файл не содержит данных листов")
        
        # Получаем размеры листов, если они есть
        sheet_dimensions = json_data.get("sheet_dimensions", {})
        
        # Создаем Excel файл с несколькими листами
        with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
            for sheet_name, sheet_data in json_data["sheets"].items():
                if sheet_data:
                    # Создаем DataFrame из данных
                    df = pd.DataFrame(sheet_data)
                    
                    # Если есть информация о размерах, проверяем соответствие
                    if sheet_name in sheet_dimensions:
                        expected_rows, expected_cols = sheet_dimensions[sheet_name]
                        # Добавляем пустые строки, если их не хватает
                        while len(df) < expected_rows:
                            df = df.append(pd.Series([None] * len(df.columns)), ignore_index=True)
                        # Добавляем пустые столбцы, если их не хватает
                        while len(df.columns) < expected_cols:
                            df[len(df.columns)] = None
                    
                    # Сохраняем лист в Excel файл
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        
        # Если есть информация о группировке, применяем ее
        if "grouping_info" in json_data and json_data["grouping_info"]:
            _apply_grouping_to_excel(output_excel_path, json_data["grouping_info"])
        
        # Если есть информация о форматировании, применяем ее
        if "formatting_info" in json_data and json_data["formatting_info"]:
            _apply_formatting_to_excel(output_excel_path, json_data["formatting_info"])
        
        return True
    except Exception as e:
        print(f"Ошибка преобразования JSON в Excel: {e}")
        return False

def _apply_grouping_to_excel(excel_path: str, grouping_info: Dict[str, Any]) -> bool:
    """
    Применяет информацию о группировке к Excel файлу.
    
    Args:
        excel_path: Путь к Excel файлу
        grouping_info: Информация о группировке
        
    Returns:
        True если успешно
    """
    try:
        from openpyxl import load_workbook
        
        # Открываем Excel файл для применения группировки
        wb = load_workbook(excel_path)
        
        # Применяем группировку для каждого листа
        for sheet_name, group_info in grouping_info.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Применяем группировку строк
                if 'row_outline_levels' in group_info:
                    for row_num, outline_level in group_info['row_outline_levels'].items():
                        # Если уровень отрицательный, строка свернута
                        abs_level = abs(outline_level)
                        if abs_level > 0:
                            ws.row_dimensions[int(row_num)].outline_level = abs_level
                            # Если исходный уровень был отрицательным, строка свернута
                            if outline_level < 0:
                                ws.row_dimensions[int(row_num)].hidden = True
                
                # Применяем группировку столбцов
                if 'column_outline_levels' in group_info:
                    for col_letter, outline_level in group_info['column_outline_levels'].items():
                        # Если уровень отрицательный, столбец свернут
                        abs_level = abs(outline_level)
                        if abs_level > 0:
                            ws.column_dimensions[col_letter].outline_level = abs_level
                            # Если исходный уровень был отрицательным, столбец свернут
                            if outline_level < 0:
                                ws.column_dimensions[col_letter].hidden = True
        
        # Сохраняем изменения
        wb.save(excel_path)
        wb.close()
        
        return True
    except Exception as e:
        print(f"Ошибка применения группировки к Excel файлу: {e}")
        return False

def _apply_formatting_to_excel(excel_path: str, formatting_info: Dict[str, Any]) -> bool:
    """
    Применяет информацию о форматировании к Excel файлу.
    
    Args:
        excel_path: Путь к Excel файлу
        formatting_info: Информация о форматировании
        
    Returns:
        True если успешно
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import column_index_from_string
        
        # Открываем Excel файл для применения форматирования
        wb = load_workbook(excel_path)
        
        # Применяем форматирование для каждого листа
        for sheet_name, format_info in formatting_info.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Применяем форматирование ячеек
                if 'cell_formats' in format_info:
                    for cell_address, cell_format in format_info['cell_formats'].items():
                        try:
                            row, col = cell_address.split('_')
                            col_index = column_index_from_string(col)
                            cell = ws.cell(row=int(row), column=col_index)
                            
                            # Применяем шрифт
                            if 'font' in cell_format:
                                font_info = cell_format['font']
                                cell.font = Font(
                                    name=font_info.get('name'),
                                    size=font_info.get('size'),
                                    bold=font_info.get('bold'),
                                    italic=font_info.get('italic'),
                                    color=font_info.get('color')
                                )
                            
                            # Применяем заливку
                            if 'fill' in cell_format:
                                fill_info = cell_format['fill']
                                if fill_info.get('fgColor'):
                                    cell.fill = PatternFill(
                                        patternType=fill_info.get('patternType'),
                                        fgColor=fill_info.get('fgColor')
                                    )
                            
                            # Применяем границы
                            if 'border' in cell_format:
                                border_info = cell_format['border']
                                sides = {}
                                for side in ['left', 'right', 'top', 'bottom']:
                                    if border_info.get(side):
                                        sides[side] = Side(style=border_info[side])
                                cell.border = Border(**sides)
                            
                            # Применяем выравнивание
                            if 'alignment' in cell_format:
                                align_info = cell_format['alignment']
                                cell.alignment = Alignment(
                                    horizontal=align_info.get('horizontal'),
                                    vertical=align_info.get('vertical')
                                )
                        except Exception as e:
                            # Пропускаем ошибки форматирования отдельных ячеек
                            continue
                
                # Применяем объединение ячеек
                if 'merged_cells' in format_info:
                    for merged_range in format_info['merged_cells']:
                        try:
                            ws.merge_cells(merged_range)
                        except Exception as e:
                            # Пропускаем ошибки объединения ячеек
                            continue
        
        # Сохраняем изменения
        wb.save(excel_path)
        wb.close()
        
        return True
    except Exception as e:
        print(f"Ошибка применения форматирования к Excel файлу: {e}")
        return False


# Пример использования
if __name__ == "__main__":
    # Пример использования класса
    print("📚 Демонстрация сохранения иерархии Excel файлов")
    print("=" * 50)
    
    # Создаем экземпляр класса (замените путь на реальный файл)
    # preserver = ExcelHierarchyPreserver("путь_к_вашему_файлу.xlsx")
    
    # Получаем информацию о структуре
    # structure_info = preserver.get_structure_info()
    # print("Структура файла:")
    # print(json.dumps(structure_info, ensure_ascii=False, indent=2))
    
    # Сохраняем в JSON
    # preserver.to_json("результаты/иерархия_файла.json")
    # print("✅ Структура сохранена в JSON")
    
    # Сохраняем в CSV
    # preserver.to_hierarchical_csv("результаты/csv_файлы")
    # print("✅ Структура сохранена в CSV файлы")
    
    print("Для использования раскомментируйте строки выше и укажите путь к реальному Excel файлу.")